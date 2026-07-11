#!/usr/bin/env python
# -*- coding: utf-8 -*-
r"""
Ruby PHR Link - Excel History Importer v1.0.0
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
역할: 과거 손으로 정리해둔 '남기동 실손보험 정리.xlsx' 데이터를 데이터베이스로 이관(마이그레이션)
주요 기능:
  - D:\남기동\수원\실손보험\남기동 실손보험 정리.xlsx 로딩
  - Sheet1의 3행부터 진료일, 병원명, 본인부담금, 상병명, 상병코드를 파싱
  - 중복 체크를 거쳐 hospital_visits 및 insurance_claims 테이블에 안전 적재
  - traceback 및 app.log 로깅 지원
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import os
import sys
import sqlite3
import logging
import traceback
import openpyxl
from pathlib import Path
from datetime import datetime

# ─── 경로 및 설정 ────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "medical_history.db"
EXCEL_PATH = Path(r"D:\남기동\수원\실손보험\남기동 실손보험 정리.xlsx")

LOG_FILE = BASE_DIR / "logs" / "app.log"
logger = logging.getLogger("RubyPHRLink.ExcelImporter")
logger.setLevel(logging.INFO)

# 파일 핸들러 (UTF-8 인코딩 명시)
if not (BASE_DIR / "logs").exists():
    (BASE_DIR / "logs").mkdir(parents=True, exist_ok=True)
file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
file_formatter = logging.Formatter('[%(asctime)s] %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
file_handler.setFormatter(file_formatter)
logger.addHandler(file_handler)

def log_info(msg):
    print(msg)
    logger.info(msg)

def log_warn(msg):
    print(f"⚠️ WARN: {msg}")
    logger.warning(msg)

def log_error(msg):
    print(f"❌ ERROR: {msg}", file=sys.stderr)
    logger.error(msg)

# ─── 날짜 포맷 정규화 헬퍼 ──────────────────────────────────────────────────────
def normalize_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    
    val_str = str(val).strip()
    # 2020.11.05 이나 2020-11-05 형태 파싱
    val_str = val_str.replace(".", "-").replace("/", "-").replace(" ", "")
    
    # YYYY-MM-DD 패턴 검출
    match = re.match(r"(\d{4})-(\d{2})-(\d{2})", val_str)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
        
    return None

import re # re 모듈 임포트 보장

# ─── 엑셀 이관 실행 ──────────────────────────────────────────────────────────
def import_excel_data():
    log_info("🔄 Ruby PHR Link 엑셀 데이터 마이그레이션 가동...")
    
    if not EXCEL_PATH.exists():
        log_error(f"엑셀 파일이 지정된 경로에 존재하지 않습니다: {EXCEL_PATH}")
        return

    if not DB_PATH.exists():
        log_error(f"데이터베이스 파일이 존재하지 않습니다. db_manager.py를 실행하십시오: {DB_PATH}")
        return

    conn = None
    try:
        # 1. 엑셀 파일 로드 (수식 계산 후 값만 가져오기)
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        if "Sheet1" not in wb.sheetnames:
            log_error("엑셀 파일 내에 'Sheet1' 시트가 존재하지 않습니다.")
            return
            
        sheet = wb["Sheet1"]
        
        # 2. SQLite 데이터베이스 연결
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("PRAGMA foreign_keys = ON;")

        success_visits = 0
        success_claims = 0
        skip_count = 0
        
        # 3행부터 순회 시작
        for r in range(3, sheet.max_row + 1):
            row_vals = [sheet.cell(r, c).value for c in range(1, 9)]
            
            # 날짜(B열, index 1)가 비어있으면 데이터가 없는 행으로 스킵
            raw_date = row_vals[1]
            if not raw_date:
                continue
                
            visit_date = normalize_date(raw_date)
            if not visit_date:
                log_warn(f"  ⚠️ {r}행: 날짜 포맷이 올바르지 않아 스킵합니다 (원문: {raw_date})")
                continue
                
            hospital_name = str(row_vals[2]).strip() if row_vals[2] else "미상 병원"
            patient_share = row_vals[3]  # 본인 부담금 (int)
            diagnosis_code = str(row_vals[4]).strip() if row_vals[4] else ""
            diagnosis_name = str(row_vals[5]).strip() if row_vals[5] else ""
            tel_number = str(row_vals[6]).strip() if row_vals[6] else ""
            
            # 3. 진단명 포맷팅 (예: "기타경추간판전위 (M502)")
            if diagnosis_name and diagnosis_code:
                diagnosis = f"{diagnosis_name} ({diagnosis_code})"
            elif diagnosis_name:
                diagnosis = diagnosis_name
            else:
                diagnosis = "미지정"
                
            memo = "엑셀 마이그레이션 적재 건"
            if tel_number:
                memo += f" (병원연락처: {tel_number})"
                
            # 4. 중복 진료기록 체크 (날짜 + 병원명 기준)
            cursor.execute("""
                SELECT visit_id FROM hospital_visits WHERE visit_date = ? AND hospital_name = ?
            """, (visit_date, hospital_name))
            row = cursor.fetchone()
            
            if row:
                visit_id = row[0]
                skip_count += 1
            else:
                # 신규 진료기록 추가
                cursor.execute("""
                    INSERT INTO hospital_visits (visit_date, hospital_name, doctor_name, diagnosis, memo)
                    VALUES (?, ?, ?, ?, ?)
                """, (visit_date, hospital_name, "미상", diagnosis, memo))
                visit_id = cursor.lastrowid
                success_visits += 1

            # 5. 실손보험 비용정보 적재 (환급/청구 관련)
            if patient_share is not None:
                # 중복 청구 레코드 체크 (visit_id 기준)
                cursor.execute("""
                    SELECT claim_id FROM insurance_claims WHERE visit_id = ? AND patient_share = ?
                """, (visit_id, patient_share))
                claim_row = cursor.fetchone()
                
                if not claim_row:
                    cursor.execute("""
                        INSERT INTO insurance_claims (visit_id, doc_id, total_cost, patient_share, non_covered_cost, claim_status, memo)
                        VALUES (?, ?, ?, ?, ?, 'CLAIMED', '엑셀 수동 이관건')
                    """, (visit_id, None, patient_share, patient_share, 0.0))
                    success_claims += 1

        conn.commit()
        log_info(f"\n📊 [엑셀 마이그레이션 결과 리포트]")
        log_info(f"  - 신규 진료기록 적재 : {success_visits} 건")
        log_info(f"  - 신규 실손보험 청구 : {success_claims} 건")
        log_info(f"  - 중복 스킵(기존진료) : {skip_count} 건")
        
        # 전체 통계 조회
        try:
            import db_manager
            db_manager.print_status()
        except ImportError:
            pass
            
    except Exception as e:
        log_error(f"엑셀 마이그레이션 실패: {e}")
        log_error(traceback.format_exc())
        if conn:
            conn.rollback()
    finally:
        if conn:
            conn.close()

if __name__ == "__main__":
    import_excel_data()
